"""
AV Deliverable Creator
======================
Streamlit app that takes an AV workbook + transcript .docx files + an optional
Drive-link CSV and produces a fully populated deliverable spreadsheet matching
the _AV_Deliverable_Creator_Template.xlsx format.

Columns populated:
  C – Asset Link (Google Drive)
  D – Transcript Book Doc Link
  E – Individual Transcript Link

Run locally:
    pip install streamlit openpyxl pandas python-docx
    streamlit run av_deliverable_creator.py
"""

import io
import re
import zipfile
from datetime import datetime, timedelta

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
BOOK_CHUNK_SIZE = 250   # rows per transcript book

EXPECTED_HEADERS = [
    "Date", "File Name", "Asset Link (Google Drive)",
    "Transcript Book Doc Link", "Individual Transcript Link",
    "Length", "James Begins Speaking", "Location/Outlet/Show",
    "Audio/Video", "Type of Event", "Source", "Link"
]

COL_DATE        = 1   # A
COL_FILENAME    = 2   # B
COL_ASSET_LINK  = 3   # C
COL_BOOK_LINK   = 4   # D
COL_TRANS_LINK  = 5   # E
COL_LENGTH      = 6   # F
COL_BEGINS      = 7   # G
COL_OUTLET      = 8   # H
COL_AV          = 9   # I
COL_EVENT_TYPE  = 10  # J
COL_SOURCE      = 11  # K
COL_PUBLIC_LINK = 12  # L

HEADER_BG   = "1F3864"   # dark navy
HEADER_FG   = "FFFFFF"
ALT_ROW_BG  = "EBF0FA"   # light blue tint
LINK_COLOR  = "1155CC"   # Google-blue hyperlink

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def normalize(name: str) -> str:
    """Strip whitespace, remove extension, lowercase for matching."""
    name = name.strip()
    name = re.sub(r'\.(mp4|mp3|docx|doc|txt)$', '', name, flags=re.IGNORECASE)
    return name.lower()


def extract_canonical_name(raw: str) -> str:
    """
    Return the best display filename from a workbook cell value.
    Handles raw strings and Excel CONCATENATE formula strings.
    """
    if not raw:
        return ""
    raw = str(raw).strip()
    # If it's a CONCATENATE formula, extract the outlet from H-column reference
    # and fall back to the raw string (openpyxl returns formula text not result)
    if raw.startswith("="):
        return raw  # will be handled separately
    return raw


def build_filename_index(drive_df: pd.DataFrame) -> dict:
    """
    Build a dict of  normalize(filename) -> drive_url
    from the uploaded Drive-link CSV.
    Expected columns: filename, drive_url  (case-insensitive)
    """
    cols = [c.lower().strip() for c in drive_df.columns]
    # Accept flexible column names
    fn_col = next((c for c in drive_df.columns if c.lower().strip() in
                   ("filename", "file name", "file_name", "name")), None)
    url_col = next((c for c in drive_df.columns if c.lower().strip() in
                    ("drive_url", "drive url", "url", "link", "drive_link")), None)
    if not fn_col or not url_col:
        return {}
    index = {}
    for _, row in drive_df.iterrows():
        fn = str(row[fn_col]).strip()
        url = str(row[url_col]).strip()
        if fn and url and url.lower() != "nan":
            index[normalize(fn)] = url
    return index


def build_book_link_map(book_df: pd.DataFrame) -> dict:
    """
    Build a dict of  book_number (1-based int) -> drive_url
    from the uploaded Transcript Books CSV.
    Expected columns: book_number (or book), drive_url
    """
    bn_col  = next((c for c in book_df.columns if c.lower().strip() in
                    ("book_number", "book number", "book", "book #", "book#")), None)
    url_col = next((c for c in book_df.columns if c.lower().strip() in
                    ("drive_url", "drive url", "url", "link", "drive_link")), None)
    if not bn_col or not url_col:
        return {}
    index = {}
    for _, row in book_df.iterrows():
        try:
            bn = int(row[bn_col])
        except (ValueError, TypeError):
            continue
        url = str(row[url_col]).strip()
        if url and url.lower() != "nan":
            index[bn] = url
    return index


def row_data_index(ws, filename_col_idx) -> list:
    """
    Return list of (excel_row_num, filename_str) for all non-empty data rows.
    Skips the header row (row 1).
    filename_col_idx: 1-based column index for filename
    """
    rows = []
    for r in range(2, ws.max_row + 1):
        cell_b = ws.cell(r, filename_col_idx).value
        if cell_b is None:
            continue
        fn = str(cell_b).strip()
        if fn:
            rows.append((r, fn))
    return rows


def hyperlink_cell(ws, row, col, url, display_text=None):
    """Write a clickable hyperlink into a cell."""
    cell = ws.cell(row, col)
    if not display_text:
        display_text = url
    cell.value = display_text
    cell.hyperlink = url
    cell.font = Font(color=LINK_COLOR, underline="single", name="Calibri", size=10)


def style_header(ws, max_col):
    """Apply header row styling."""
    for c in range(1, max_col + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color=HEADER_FG, name="Calibri", size=10)
        cell.fill = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_data_rows(ws, data_rows_excel, max_col):
    """Zebra-stripe data rows, center-align, set row height."""
    thin = Side(style="thin", color="D0D7E8")
    border = Border(bottom=thin)
    for idx, (r, _) in enumerate(data_rows_excel):
        fill_color = ALT_ROW_BG if idx % 2 == 1 else "FFFFFF"
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if cell.hyperlink:
                continue  # don't override hyperlink style
            cell.fill = PatternFill("solid", start_color=fill_color)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = border
        ws.row_dimensions[r].height = 16


def set_column_widths(ws):
    widths = {
        COL_DATE:       12,
        COL_FILENAME:   38,
        COL_ASSET_LINK: 22,
        COL_BOOK_LINK:  22,
        COL_TRANS_LINK: 22,
        COL_LENGTH:     10,
        COL_BEGINS:     14,
        COL_OUTLET:     26,
        COL_AV:         10,
        COL_EVENT_TYPE: 14,
        COL_SOURCE:     12,
        COL_PUBLIC_LINK:32,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ─────────────────────────────────────────────
# CORE PROCESSING
# ─────────────────────────────────────────────

def process_workbook(
    wb_bytes: bytes,
    transcript_files: dict,       # {normalize(stem): (display_name, bytes)}
    asset_link_index: dict,       # normalize(filename) -> drive_url
    transcript_link_index: dict,  # normalize(stem) -> drive_url
    book_link_map: dict,          # book_number -> drive_url
    book_chunk_size: int,
    candidate_prefix: str,
    column_mapping: dict,         # {'asset': colname, 'book': colname, 'transcript': colname}
) -> tuple[bytes, dict]:
    """
    Main processing function.
    Returns (output_xlsx_bytes, stats_dict).
    """
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
    ws = wb.active

    # Map column names to 1-based indices
    header_row = [str(ws.cell(1, c+1).value).strip() for c in range(ws.max_column)]
    def col_idx(colname):
        return header_row.index(colname) + 1 if colname in header_row else None
    asset_col_idx = col_idx(column_mapping['asset'])
    book_col_idx = col_idx(column_mapping['book'])
    trans_col_idx = col_idx(column_mapping['transcript'])
    filename_col_idx = col_idx('File Name') or 2  # fallback to 2 if not found

    data_rows = row_data_index(ws, filename_col_idx)
    stats = {
        "total_rows": len(data_rows),
        "asset_links_filled": 0,
        "transcript_links_filled": 0,
        "book_links_filled": 0,
        "missing_asset": [],
        "missing_transcript": [],
    }

    for data_idx, (excel_row, raw_filename) in enumerate(data_rows):
        # ── determine canonical stem ──────────────────────────────────────
        # If the filename cell is a formula, try to build it from date + outlet
        if raw_filename.startswith("="):
            date_val = ws.cell(excel_row, COL_DATE).value
            outlet   = ws.cell(excel_row, COL_OUTLET).value or ""
            if date_val and isinstance(date_val, datetime):
                prefix = date_val.strftime("%y%m%d")
            else:
                prefix = ""
            stem = f"{prefix} {candidate_prefix} on {outlet.strip()}".strip()
        else:
            stem = raw_filename

        norm_stem = normalize(stem)

        # ── Individual Transcript Link (robust: .docx/.txt) ──
        trans_url = (
            transcript_link_index.get(norm_stem)
            or transcript_link_index.get(norm_stem + ".docx")
            or transcript_link_index.get(norm_stem + ".txt")
        )
        if not trans_url and norm_stem in transcript_files:
            trans_url = None
        if trans_url and trans_col_idx:
            # Use the correct extension for display
            if transcript_link_index.get(norm_stem + ".txt"):
                display_name = stem.strip() + ".txt"
            else:
                display_name = stem.strip() + ".docx"
            hyperlink_cell(ws, excel_row, trans_col_idx, trans_url, display_name)
            stats["transcript_links_filled"] += 1
        else:
            stats["missing_transcript"].append(stem)

        # ── Transcript Book Doc Link (user-mapped column) ──
        book_number = (data_idx // book_chunk_size) + 1
        book_url = book_link_map.get(book_number)
        if book_url and book_col_idx:
            hyperlink_cell(ws, excel_row, book_col_idx, book_url, f"Transcript Book {book_number}")
            stats["book_links_filled"] += 1
        elif book_col_idx:
            ws.cell(excel_row, book_col_idx).value = f"[Book {book_number} — paste link]"
            ws.cell(excel_row, book_col_idx).font = Font(
                color="999999", italic=True, name="Calibri", size=10
            )

    # ── Styling ──────────────────────────────────────────────────────────
    style_header(ws, ws.max_column)
    style_data_rows(ws, data_rows, ws.max_column)
    set_column_widths(ws)
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    print("Processing complete. Stats:", stats)
    return out.getvalue(), stats


# ─────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="AV Deliverable Creator",
    page_icon="📋",
    layout="wide",
)

st.markdown("""
<style>
    .main-title { font-size: 2rem; font-weight: 700; color: #1F3864; margin-bottom: 0; }
    .sub-title  { font-size: 1rem; color: #666; margin-bottom: 1.5rem; }
    .section-header { font-size: 1.1rem; font-weight: 600; color: #1F3864;
                       border-left: 4px solid #1F3864; padding-left: 10px; margin: 1rem 0 0.5rem; }
    .stat-box { background: #EBF0FA; border-radius: 8px; padding: 12px 18px;
                text-align: center; border: 1px solid #C5D3ED; }
    .stat-num { font-size: 1.8rem; font-weight: 700; color: #1F3864; }
    .stat-label { font-size: 0.8rem; color: #555; }
    .tip { background: #FFF8E1; border-left: 4px solid #F9A825;
           padding: 8px 14px; border-radius: 4px; font-size: 0.85rem; color: #5D4037; }
</style>
""", unsafe_allow_html=True)

st.markdown('<p class="main-title">📋 AV Deliverable Creator</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">Populate Drive links, transcript links, and book assignments in your AV tracking spreadsheet.</p>', unsafe_allow_html=True)

# ── STEP 1: Workbook ─────────────────────────────────────────────────────────
st.markdown('<p class="section-header">Step 1 — Upload AV Workbook</p>', unsafe_allow_html=True)

def validate_workbook_headers(wb_bytes: bytes, expected_headers: list[str]) -> tuple[bool, list[str]]:
    """Check if the uploaded workbook has the expected headers."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
        ws = wb.active
        headers = [str(ws.cell(1, c+1).value).strip() for c in range(len(expected_headers))]
        return headers == expected_headers, headers
    except Exception:
        return False, []


workbook_file = st.file_uploader(
    "Upload your AV tracking spreadsheet (.xlsx)",
    type=["xlsx"],
    key="workbook"
)

candidate_prefix = st.text_input(
    "Candidate identifier used in filenames",
    value="MI James",
    help='The short name that appears in every filename, e.g. "MI James" or "OH Brown"'
)

# ── COLUMN MAPPING ──
column_mapping = {}
headers = []
if workbook_file:
    file_bytes = workbook_file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        headers = [str(ws.cell(1, c+1).value).strip() for c in range(ws.max_column)]
    except Exception:
        st.error("Could not read headers from uploaded workbook.")
        st.stop()

    st.markdown('<p class="section-header">Step 1a — Map Columns</p>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        asset_col = st.selectbox("Asset Link (Google Drive) column", headers, key="asset_col")
    with col2:
        book_col = st.selectbox("Transcript Book Doc Link column", headers, key="book_col")
    with col3:
        trans_col = st.selectbox("Individual Transcript Link column", headers, key="trans_col")
    column_mapping = {
        "asset": asset_col,
        "book": book_col,
        "transcript": trans_col,
    }



# ── STEP 2: Asset, Transcript, and Book Link CSVs ───────────────────────────
col1, col2, col3 = st.columns(3)
with col1:
    asset_csv = st.file_uploader(
        "Asset Links CSV\n(filename → Drive URL for .mp4/.mp3 files)",
        type=["csv"],
        key="asset_csv"
    )
with col2:
    transcript_csv = st.file_uploader(
        "Individual Transcript Links CSV\n(filename → Drive URL for .docx/.txt files)",
        type=["csv"],
        key="transcript_csv"
    )
with col3:
    book_csv = st.file_uploader(
        "Transcript Book Links CSV\n(book_number → Drive URL)",
        type=["csv"],
        key="book_csv"
    )

# ── STEP 3: Transcript .docx files ──────────────────────────────────────────
st.markdown('<p class="section-header">Step 3 — Upload Transcript .docx Files (optional)</p>', unsafe_allow_html=True)
transcript_docx_files = st.file_uploader(
    "Upload individual transcript Word documents",
    type=["docx"],
    accept_multiple_files=True,
    key="transcripts"
)

# ── STEP 4: Book chunk size ──────────────────────────────────────────────────
st.markdown('<p class="section-header">Step 4 — Transcript Book Settings</p>', unsafe_allow_html=True)
book_chunk = st.number_input(
    "Rows per transcript book",
    min_value=10,
    max_value=1000,
    value=BOOK_CHUNK_SIZE,
    step=10,
    help="Each book covers this many consecutive rows. Default is 250."
)


# ── OUTPUT FOLDER & WORD FILE OPTIONS ───────────────────────────────────────
st.divider()
st.markdown('<p class="section-header">Step 5 — Output Options</p>', unsafe_allow_html=True)
from pathlib import Path
output_folder = st.text_input(
    "Output folder for Word files (absolute path)",
    value=str((Path(__file__).parent / "output").resolve()),
    help="Where to save uploaded and/or generated Word files. Must exist."
)
save_uploaded = st.checkbox("Save uploaded transcript .docx files to output folder", value=True)
generate_word = st.checkbox("Generate new Word files for each row (from spreadsheet/CSV)", value=False)

# ── PROCESS BUTTON ────────────────────────────────────────────────────────────


if st.button("▶ Generate Deliverable Spreadsheet", type="primary", use_container_width=True):
    if not workbook_file:
        st.error("Please upload the AV workbook first.")
        st.stop()
    if not column_mapping or not all(column_mapping.values()):
        st.error("Please map all required columns before proceeding.")
        st.stop()
    # file_bytes already read above

    with st.spinner("Processing..."):
        # Build asset link index
        asset_index = {}
        if asset_csv:
            try:
                df_a = pd.read_csv(asset_csv)
                asset_index = build_filename_index(df_a)
                st.success(f"✅ Loaded {len(asset_index)} asset link entries.")
            except Exception as e:
                st.warning(f"Could not parse asset CSV: {e}. Please ensure it has columns 'filename' and 'drive_url'.")

        # Build transcript link index
        trans_index = {}
        if transcript_csv:
            try:
                df_t = pd.read_csv(transcript_csv)
                trans_index = build_filename_index(df_t)
                st.success(f"✅ Loaded {len(trans_index)} transcript link entries.")
            except Exception as e:
                st.warning(f"Could not parse transcript CSV: {e}. Please ensure it has columns 'filename' and 'drive_url'.")

        # Build book link map
        book_map = {}
        if book_csv:
            try:
                df_b = pd.read_csv(book_csv)
                book_map = build_book_link_map(df_b)
                st.success(f"✅ Loaded {len(book_map)} transcript book link entries.")
            except Exception as e:
                st.warning(f"Could not parse book CSV: {e}. Please ensure it has columns 'book_number' and 'drive_url'.")

        # Build uploaded transcript docx index and save if requested
        uploaded_transcripts = {}
        output_path = Path(output_folder)
        if save_uploaded and output_path.exists():
            for f in (transcript_docx_files or []):
                key = normalize(f.name)
                data = f.read()
                uploaded_transcripts[key] = (f.name, data)
                # Save uploaded file
                out_file = output_path / f.name
                with open(out_file, "wb") as out_f:
                    out_f.write(data)
        else:
            for f in (transcript_docx_files or []):
                key = normalize(f.name)
                uploaded_transcripts[key] = (f.name, f.read())

        # Optionally generate new Word files for each row
        if generate_word and output_path.exists():
            import docx
            # Example: generate a Word file for each row in the spreadsheet
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
                ws = wb.active
                header_row = [str(ws.cell(1, c+1).value).strip() for c in range(ws.max_column)]
                filename_col_idx = header_row.index('File Name') + 1 if 'File Name' in header_row else 2
                for r in range(2, ws.max_row + 1):
                    file_name = str(ws.cell(r, filename_col_idx).value).strip()
                    if not file_name:
                        continue
                    doc = docx.Document()
                    doc.add_heading(file_name, 0)
                    # Optionally add more content from the row here
                    out_file = output_path / (file_name + ".docx")
                    doc.save(str(out_file))
            except Exception as e:
                st.warning(f"Could not generate Word files: {e}")

        # Run processing
        try:
            output_bytes, stats = process_workbook(
                wb_bytes=file_bytes,
                transcript_files=uploaded_transcripts,
                asset_link_index=asset_index,
                transcript_link_index=trans_index,
                book_link_map=book_map,
                book_chunk_size=int(book_chunk),
                candidate_prefix=candidate_prefix.strip(),
                column_mapping=column_mapping,
            )
        except Exception as e:
            st.error(f"Processing error: {e}")
            st.exception(e)
            st.stop()

    # ── RESULTS ──────────────────────────────────────────────────────────────
    st.success("✅ Spreadsheet generated successfully!")

    # Preview first 5 rows of output (as DataFrame)
    try:
        preview_df = pd.read_excel(io.BytesIO(output_bytes))
        st.markdown("#### Preview of Output (first 5 rows):")
        st.dataframe(preview_df.head(5), use_container_width=True)
    except Exception as e:
        st.warning(f"Could not preview output: {e}")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f'<div class="stat-box"><div class="stat-num">{stats["total_rows"]}</div>'
                    f'<div class="stat-label">Total Rows</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="stat-box"><div class="stat-num">{stats["transcript_links_filled"]}</div>'
                    f'<div class="stat-label">Transcript Links Filled</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="stat-box"><div class="stat-num">{stats["book_links_filled"]}</div>'
                    f'<div class="stat-label">Book Links Filled</div></div>', unsafe_allow_html=True)

    st.markdown("### ⚠️ Missing Transcript Links")
    if stats["missing_transcript"]:
        for name in stats["missing_transcript"]:
            st.write(f"• {name}")
    else:
        st.write("None!")

    # ── DOWNLOAD ─────────────────────────────────────────────────────────────
    st.download_button(
        label="⬇️ Download Completed Spreadsheet",
        data=output_bytes,
        file_name=f"AV_Deliverable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )

# ── SIDEBAR: CSV format guides ────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📄 CSV Format Guide")
    st.markdown("**Asset / Transcript Links CSV:**")
    st.code("filename,drive_url\n250417 MI James on The Benny Show.mp4,https://drive.google.com/...\n250424 MI James on The Detroit News.mp4,https://drive.google.com/...", language="text")
    st.markdown("**Transcript Books CSV:**")
    st.code("book_number,drive_url\n1,https://drive.google.com/...\n2,https://drive.google.com/...", language="text")
    st.divider()
    st.markdown("### ℹ️ Column Reference")
    st.markdown("""
| Col | Header |
|-----|--------|
| C | Asset Link (Drive) |
| D | Transcript Book Link |
| E | Individual Transcript |
""")
    st.markdown(f"Rows per book: **{BOOK_CHUNK_SIZE}** (adjustable in Step 4)")
