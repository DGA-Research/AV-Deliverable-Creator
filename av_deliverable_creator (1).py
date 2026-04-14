"""
AV Deliverable Creator
======================
Streamlit app that takes an AV workbook + CSVs and produces a fully
populated deliverable spreadsheet.

Fixes applied:
  1. Fuzzy column matching — tolerates different column name casing/wording
  2. Formula cell evaluation — handles =CONCATENATE(...) File Name cells
  3. Asset link write — actually populates Drive Link column (was missing)
  4. Hardcoded book links for Treg Taylor (13 books)

Run locally:
    pip install streamlit openpyxl pandas python-docx
    streamlit run av_deliverable_creator.py
"""

import io
import re
from datetime import datetime

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
BOOK_CHUNK_SIZE = 31  # rows per transcript book (395 rows / 13 books)

HEADER_BG  = "1F3864"
HEADER_FG  = "FFFFFF"
ALT_ROW_BG = "EBF0FA"
LINK_COLOR = "1155CC"

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def normalize(name: str) -> str:
    """Strip whitespace, remove extension, lowercase for matching."""
    name = name.strip()
    name = re.sub(r'\.(mp4|mp3|docx|doc|txt)$', '', name, flags=re.IGNORECASE)
    return name.lower()


def fuzzy_col_match(headers: list, target: str) -> int | None:
    """
    Find 1-based column index by fuzzy matching:
    - exact match first
    - then normalize both sides (lowercase, strip spaces/punctuation)
    Returns None if not found.
    """
    norm_target = re.sub(r'[^a-z0-9]', '', target.lower())
    for i, h in enumerate(headers):
        if h == target:
            return i + 1
        if re.sub(r'[^a-z0-9]', '', h.lower()) == norm_target:
            return i + 1
    # partial match fallback
    for i, h in enumerate(headers):
        if norm_target in re.sub(r'[^a-z0-9]', '', h.lower()):
            return i + 1
    return None


def build_filename_index(df: pd.DataFrame) -> dict:
    """Build normalize(filename) -> drive_url from a CSV dataframe."""
    fn_col  = next((c for c in df.columns if c.lower().strip() in
                    ("filename", "file name", "file_name", "name")), None)
    url_col = next((c for c in df.columns if c.lower().strip() in
                    ("drive_url", "drive url", "url", "link", "drive_link")), None)
    if not fn_col or not url_col:
        return {}
    index = {}
    for _, row in df.iterrows():
        fn  = str(row[fn_col]).strip()
        url = str(row[url_col]).strip()
        if fn and url and url.lower() != "nan":
            index[normalize(fn)] = url
    return index


def build_book_link_map(df: pd.DataFrame) -> dict:
    """Build book_number (int) -> drive_url from a CSV dataframe."""
    bn_col  = next((c for c in df.columns if c.lower().strip() in
                    ("book_number", "book number", "book", "book #", "book#")), None)
    url_col = next((c for c in df.columns if c.lower().strip() in
                    ("drive_url", "drive url", "url", "link", "drive_link")), None)
    if not bn_col or not url_col:
        return {}
    index = {}
    for _, row in df.iterrows():
        try:
            bn = int(row[bn_col])
        except (ValueError, TypeError):
            continue
        url = str(row[url_col]).strip()
        if url and url.lower() != "nan":
            index[bn] = url
    return index


def resolve_stem(raw_filename: str, date_val, outlet_val: str,
                 candidate_prefix: str) -> str:
    """
    Return the canonical filename stem for matching against CSVs.
    Handles both plain strings and =CONCATENATE(...) formula cells.
    """
    if not raw_filename or not str(raw_filename).startswith("="):
        return str(raw_filename).strip()

    # Formula cell — reconstruct from date + outlet
    if date_val and isinstance(date_val, datetime):
        prefix = date_val.strftime("%y%m%d")
    else:
        prefix = ""
    outlet = str(outlet_val).strip() if outlet_val else ""
    return f"{prefix} {candidate_prefix} on {outlet}".strip()


def hyperlink_cell(ws, row, col, url, display_text=None):
    cell = ws.cell(row, col)
    cell.value = display_text or url
    cell.hyperlink = url
    cell.font = Font(color=LINK_COLOR, underline="single", name="Calibri", size=10)


def style_header(ws, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color=HEADER_FG, name="Calibri", size=10)
        cell.fill = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_data_rows(ws, data_rows, max_col):
    thin   = Side(style="thin", color="D0D7E8")
    border = Border(bottom=thin)
    for idx, (r, _) in enumerate(data_rows):
        fill_color = ALT_ROW_BG if idx % 2 == 1 else "FFFFFF"
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if cell.hyperlink:
                continue
            cell.fill      = PatternFill("solid", start_color=fill_color)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = border
        ws.row_dimensions[r].height = 16


def set_column_widths(ws, col_map):
    widths = {
        col_map.get("date",        1): 12,
        col_map.get("filename",    2): 38,
        col_map.get("asset",       3): 22,
        col_map.get("transcript",  4): 22,
        col_map.get("book",        5): 22,
        col_map.get("length",      6): 10,
        col_map.get("outlet",      7): 26,
    }
    for col, w in widths.items():
        if col:
            ws.column_dimensions[get_column_letter(col)].width = w


# ─────────────────────────────────────────────
# CORE PROCESSING
# ─────────────────────────────────────────────

def process_workbook(
    wb_bytes: bytes,
    asset_link_index: dict,
    transcript_link_index: dict,
    book_link_map: dict,
    book_chunk_size: int,
    candidate_prefix: str,
) -> tuple[bytes, dict]:

    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
    ws = wb.active

    # ── Read headers & fuzzy-map columns ─────────────────────────────────
    headers = [str(ws.cell(1, c + 1).value or "").strip()
               for c in range(ws.max_column)]

    def ci(name):
        return fuzzy_col_match(headers, name)

    filename_col_idx    = ci("File Name")    or 2
    date_col_idx        = ci("Date")         or 1
    outlet_col_idx      = ci("Location/Outlet/Show") or 7
    asset_col_idx       = ci("Drive Link")   or ci("Asset Link (Google Drive)") or 3
    trans_col_idx       = ci("Individual Transcript Link") or 4
    book_col_idx        = ci("Transcript Book Link") or ci("Transcript Book Doc Link") or 5

    # Collect data rows (skip header)
    data_rows = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, filename_col_idx).value
        if val is not None and str(val).strip():
            data_rows.append((r, str(val).strip()))

    stats = {
        "total_rows":              len(data_rows),
        "asset_links_filled":      0,
        "transcript_links_filled": 0,
        "book_links_filled":       0,
        "missing_asset":           [],
        "missing_transcript":      [],
    }

    for data_idx, (excel_row, raw_filename) in enumerate(data_rows):
        date_val   = ws.cell(excel_row, date_col_idx).value
        outlet_val = ws.cell(excel_row, outlet_col_idx).value

        stem      = resolve_stem(raw_filename, date_val, outlet_val, candidate_prefix)
        norm_stem = normalize(stem)

        # ── Asset / Drive Link ────────────────────────────────────────────
        asset_url = (
            asset_link_index.get(norm_stem)
            or asset_link_index.get(norm_stem + ".mp4")
            or asset_link_index.get(norm_stem + ".mp3")
        )
        if asset_url and asset_col_idx:
            hyperlink_cell(ws, excel_row, asset_col_idx, asset_url, stem)
            stats["asset_links_filled"] += 1
        else:
            stats["missing_asset"].append(stem)

        # ── Individual Transcript Link ────────────────────────────────────
        trans_url = (
            transcript_link_index.get(norm_stem)
            or transcript_link_index.get(norm_stem + ".docx")
            or transcript_link_index.get(norm_stem + ".txt")
        )
        if trans_url and trans_col_idx:
            ext = ".txt" if transcript_link_index.get(norm_stem + ".txt") else ".docx"
            hyperlink_cell(ws, excel_row, trans_col_idx, trans_url, stem + ext)
            stats["transcript_links_filled"] += 1
        else:
            stats["missing_transcript"].append(stem)

        # ── Transcript Book Link ──────────────────────────────────────────
        book_number = (data_idx // book_chunk_size) + 1
        book_url    = book_link_map.get(book_number)
        if book_url and book_col_idx:
            hyperlink_cell(ws, excel_row, book_col_idx, book_url,
                           f"Transcript Book {book_number}")
            stats["book_links_filled"] += 1
        elif book_col_idx:
            ws.cell(excel_row, book_col_idx).value = f"[Book {book_number} — paste link]"
            ws.cell(excel_row, book_col_idx).font  = Font(
                color="999999", italic=True, name="Calibri", size=10)

    # ── Styling ───────────────────────────────────────────────────────────
    style_header(ws, ws.max_column)
    style_data_rows(ws, data_rows, ws.max_column)
    col_map = {
        "date": date_col_idx, "filename": filename_col_idx,
        "asset": asset_col_idx, "transcript": trans_col_idx,
        "book": book_col_idx, "outlet": outlet_col_idx,
    }
    set_column_widths(ws, col_map)
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), stats


# ─────────────────────────────────────────────
# HARDCODED BOOK LINKS (Treg Taylor)
# ─────────────────────────────────────────────
TREG_TAYLOR_BOOK_LINKS = {
    1:  "https://docs.google.com/document/d/10-j1FBKhPGIq2CiDJgiD8HF5TRmWjc0P/edit",
    2:  "https://docs.google.com/document/d/1GRte6HEURLpYBle8EPUgU86obQout3ew/edit",
    3:  "https://docs.google.com/document/d/1SsxHg0sZDNO0tmPpbcUiivXwI1zK6RAC/edit",
    4:  "https://docs.google.com/document/d/1zNs1h7PcHvw4L_lzs5R99_ofE2KfAxUJ/edit",
    5:  "https://docs.google.com/document/d/1ffsX7YnZwT2BbMkRyxpUef_yNobIP4LG/edit",
    6:  "https://docs.google.com/document/d/1Pj9yQ9Vst22m1sgsYx05IQdpVP-x5fck/edit",
    7:  "https://docs.google.com/document/d/1MU-27DTkEMeqmaug8Mdopt-Z5W7eGhnU/edit",
    8:  "https://docs.google.com/document/d/1v6wSR-Ff2fvbeKL-JA5rUy4ExPJy-8q7/edit",
    9:  "https://docs.google.com/document/d/1-WwEdHxLZQLGnDQSJNyPplhQMHniQ2H9/edit",
    10: "https://docs.google.com/document/d/1WuDLOwYJiJToliiMQrKBVqM7yu08LyrZ/edit",
    11: "https://docs.google.com/document/d/183e-HlTsfM7sLmB7P7vvyIS9IXwXApTN/edit",
    12: "https://docs.google.com/document/d/1K4pkbBAEVy_2ap-EyqTlSDxdDOSfo_PL/edit",
    13: "https://docs.google.com/document/d/14cWaW1FaBPGaUFsIViKEjrYrZco10oLE/edit",
}


# ─────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="AV Deliverable Creator",
    page_icon="📋",
    layout="wide",
)

st.markdown("""
<style>
    .main-title { font-size: 2rem; font-weight: 700; color: #1F3864; margin-bottom: 0; }
    .sub-title  { font-size: 1rem; color: #666; margin-bottom: 1.5rem; }
    .section-header { font-size: 1.1rem; font-weight: 600; color: #1F3864;
                       border-left: 4px solid #1F3864; padding-left: 10px; margin: 1rem 0 0.5rem; }
    .stat-box { background: #EBF0FA; border-radius: 8px; padding: 12px 18px;
                text-align: center; border: 1px solid #C5D3ED; }
    .stat-num { font-size: 1.8rem; font-weight: 700; color: #1F3864; }
    .stat-label { font-size: 0.8rem; color: #555; }
</style>
""", unsafe_allow_html=True)

st.markdown('<p class="main-title">📋 AV Deliverable Creator</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">Populate Drive links, transcript links, and book assignments in your AV tracking spreadsheet.</p>', unsafe_allow_html=True)

# ── Step 1: Workbook ──────────────────────────────────────────────────────────
st.markdown('<p class="section-header">Step 1 — Upload AV Workbook</p>', unsafe_allow_html=True)
workbook_file = st.file_uploader("Upload your AV tracking spreadsheet (.xlsx)", type=["xlsx"])

candidate_prefix = st.text_input(
    "Candidate identifier used in filenames",
    value="AK Taylor",
    help='Short name in every filename, e.g. "AK Taylor" or "MI James"'
)

# ── Step 2: CSVs ─────────────────────────────────────────────────────────────
st.markdown('<p class="section-header">Step 2 — Upload Link CSVs</p>', unsafe_allow_html=True)
col1, col2, col3 = st.columns(3)
with col1:
    asset_csv = st.file_uploader(
        "Asset Links CSV\n(filename → Drive URL for .mp4/.mp3)",
        type=["csv"], key="asset_csv"
    )
with col2:
    transcript_csv = st.file_uploader(
        "Individual Transcript Links CSV\n(filename → Drive URL for .docx/.txt)",
        type=["csv"], key="transcript_csv"
    )
with col3:
    book_csv = st.file_uploader(
        "Transcript Book Links CSV\n(book_number → Drive URL) — optional if using Treg Taylor defaults",
        type=["csv"], key="book_csv"
    )

# ── Step 3: Book settings ─────────────────────────────────────────────────────
st.markdown('<p class="section-header">Step 3 — Transcript Book Settings</p>', unsafe_allow_html=True)
use_treg_books = st.checkbox("Use Treg Taylor book links (books 1–13, 31 rows each)", value=True)
book_chunk = st.number_input(
    "Rows per transcript book",
    min_value=1, max_value=1000, value=BOOK_CHUNK_SIZE, step=1,
    help="How many consecutive rows go into each book. 395 rows ÷ 13 books = 31."
)

# ── Process ───────────────────────────────────────────────────────────────────
if st.button("▶ Generate Deliverable Spreadsheet", type="primary", use_container_width=True):
    if not workbook_file:
        st.error("Please upload the AV workbook first.")
        st.stop()

    file_bytes = workbook_file.read()

    with st.spinner("Processing..."):
        # Asset index
        asset_index = {}
        if asset_csv:
            try:
                asset_index = build_filename_index(pd.read_csv(asset_csv))
                st.success(f"✅ Loaded {len(asset_index)} asset link entries.")
            except Exception as e:
                st.warning(f"Could not parse asset CSV: {e}")

        # Transcript index
        trans_index = {}
        if transcript_csv:
            try:
                trans_index = build_filename_index(pd.read_csv(transcript_csv))
                st.success(f"✅ Loaded {len(trans_index)} transcript link entries.")
            except Exception as e:
                st.warning(f"Could not parse transcript CSV: {e}")

        # Book map
        if use_treg_books:
            book_map = TREG_TAYLOR_BOOK_LINKS
            st.success("✅ Using Treg Taylor book links (books 1–13).")
        elif book_csv:
            try:
                book_map = build_book_link_map(pd.read_csv(book_csv))
                st.success(f"✅ Loaded {len(book_map)} book link entries.")
            except Exception as e:
                st.warning(f"Could not parse book CSV: {e}")
                book_map = {}
        else:
            book_map = {}
            st.warning("No book links provided.")

        try:
            output_bytes, stats = process_workbook(
                wb_bytes=file_bytes,
                asset_link_index=asset_index,
                transcript_link_index=trans_index,
                book_link_map=book_map,
                book_chunk_size=int(book_chunk),
                candidate_prefix=candidate_prefix.strip(),
            )
        except Exception as e:
            st.error(f"Processing error: {e}")
            st.exception(e)
            st.stop()

    st.success("✅ Spreadsheet generated!")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="stat-box"><div class="stat-num">{stats["total_rows"]}</div>'
                    f'<div class="stat-label">Total Rows</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="stat-box"><div class="stat-num">{stats["asset_links_filled"]}</div>'
                    f'<div class="stat-label">Asset Links Filled</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="stat-box"><div class="stat-num">{stats["transcript_links_filled"]}</div>'
                    f'<div class="stat-label">Transcript Links Filled</div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="stat-box"><div class="stat-num">{stats["book_links_filled"]}</div>'
                    f'<div class="stat-label">Book Links Filled</div></div>', unsafe_allow_html=True)

    if stats["missing_asset"]:
        with st.expander(f"⚠️ {len(stats['missing_asset'])} missing asset links"):
            for name in stats["missing_asset"]:
                st.write(f"• {name}")

    if stats["missing_transcript"]:
        with st.expander(f"⚠️ {len(stats['missing_transcript'])} missing transcript links"):
            for name in stats["missing_transcript"]:
                st.write(f"• {name}")

    st.download_button(
        label="⬇️ Download Completed Spreadsheet",
        data=output_bytes,
        file_name=f"AV_Deliverable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📄 CSV Format Guide")
    st.markdown("**Asset / Transcript Links CSV:**")
    st.code("filename,drive_url\n250923 AK Taylor on Alaskasnewssource.mp4,https://drive.google.com/...", language="text")
    st.markdown("**Transcript Books CSV (if not using defaults):**")
    st.code("book_number,drive_url\n1,https://docs.google.com/...\n2,https://docs.google.com/...", language="text")
    st.divider()
    st.markdown("### ℹ️ Column Reference")
    st.markdown("""
| Col | Header |
|-----|--------|
| C | Drive Link |
| D | Individual Transcript Link |
| E | Transcript Book Link |
""")
    st.markdown(f"Default rows per book: **{BOOK_CHUNK_SIZE}**")
